import json
import random
import string
import threading
from datetime import datetime

from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordResetForm
from django.contrib.auth.tokens import default_token_generator
from django.contrib.auth.views import PasswordResetCompleteView, PasswordResetConfirmView
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.mail import send_mail, EmailMessage
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models.fields.files import ImageFieldFile
from django.forms import model_to_dict
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.utils.encoding import force_bytes
from django.utils.http import urlsafe_base64_encode
from openpyxl import load_workbook

from .decorators import staff_required
from .forms import *
from .models import Partie


def index_view(request):
    user = request.user

    if user.is_authenticated:
        if user.is_staff:
            return redirect("admin_")
        else:
            return redirect("Employe")

    else:
        return render(request, "app1/index.html")


# LOGIN ########################################################
def login_view(request):
    data = dict()
    if request.method == 'POST':
        form = LoginForm(request.POST)

        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')

            user = authenticate(username=username, password=password)

            if user and user.is_active:
                data["valid"] = True
                login(request, user)
                if user.is_staff:
                    data["staff"] = True
                else:
                    data["staff"] = False
            else:
                data["valid"] = False
        return JsonResponse(data)

    return render(request, "app1/login.html")


# DECONNEXION #######################################################
def logout_view(request):
    logout(request)
    return redirect("index")


# PAGE ADMIN ########################################################
@staff_required
def admin_program_view(request):
    parties = Partie.objects.all()
    chapitres = Chapitre.objects.all()
    articles = Article.objects.all()
    year = datetime.datetime.now().year
    now = datetime.datetime.now()
    annonces = Annonce.objects.filter(Q(date_fin__gte=now))
    annonces_ = Annonce.objects.filter(date_fin__gte=now,inscription=None)

    demandes = set()
    for demande in Demande.objects.all():
        if demande.etat == "acceptée":
            demandes.add(demande.id)

    articles_ = Article.objects.filter(demande__in=demandes) | Article.objects.filter(demande=None)
    pvs = Pv.objects.all()
    non_read = Message.objects.filter(is_read=False).count()
    messages = Message.objects.filter(receiver=request.user)


    context = {"parties": parties, "chapitres": chapitres, "articles": articles, "articles_": articles_, "annonces": annonces,
               "annonces_": annonces_, "pvs": pvs, "non_read": non_read, "messages": messages}
    if Budget.objects.all().count() != 0:
        budget_object = Budget.objects.get(date__year=year)
        context["solde"] = budget_object.solde
        context["budget_object"] = budget_object


    return render(request, "app1/admin_finale.html", context)

def partie_create_ajax_view(request):
    data = dict()

    if request.GET.get("titre") is not None:
        titre = request.GET.get("titre")

        if Partie.objects.filter(titre=titre).exists():
            data["message_titre"] = "Ce nom existe déja!"
            data["error"] = True

    """
    if request.GET.get("budget") is not None:
        budget = int(request.GET.get("budget"))
        solde = Budget.objects.first().solde
        solde -= budget
        
        if solde < 0:
            data["solde"] = Budget.objects.first().solde
            data["error"] =True
            data["message_budget"] = "Budget dépasse le budget total restant"
        else:
            data["solde"] = solde
    """

    return JsonResponse(data)

# AJOUT D'UNE PARTIE ########################################################
def partie_create_view(request):
    demandes = Demande.objects.all()
    year = datetime.datetime.now().year
    budget_object = Budget.objects.filter(date__year=year).first()
    solde = budget_object.solde
    context = {"solde": solde, "demandes": demandes}
    data = dict()

    if request.method == "POST":
        form = PartieModelForm(request.POST)

        if form.is_valid():
            budget = form.cleaned_data["budget"]
            partie = Partie.objects.create(budget=budget, titre=request.POST.get("titre"), budget_restant=budget
                                           ,solde=budget)
            partie.save()
            data["partie"] = model_to_dict(partie)
            budget_object.solde -= budget
            budget_object.debit += budget
            budget_object.save()
            data["message"] = "Partie ajoutée"
            data["form_valid"] = True
            data["solde"] = budget_object.solde
        else:
            field_errors = {}
            valid_fields = []
            for field in form:
                if field.errors:
                    field_errors[field.name] = field.errors
                    continue
                else:
                    valid_fields.append(field.name)

            data["form_valid"] = False
            data["field_errors"] = field_errors
            data["valid_fields"] = valid_fields

        return JsonResponse(data)

    return render(request, "app1/AjouterPartie.html", context)

def partie_update_ajax_view(request):
    partie = model_to_dict(Partie.objects.get(pk=request.GET.get('partie')))
    data = {"partie": partie}

    return JsonResponse(data)


# MISE A JOUR D'UNE PARTIE ######################################################
def partie_update_view(request):
    parties = Partie.objects.all()
    year = datetime.datetime.now().year
    budget_object = Budget.objects.filter(date__year=year).first()
    data = dict()

    if request.method == "POST":
        partie = Partie.objects.get(pk=request.POST.get('partie'))
        old_budget = partie.budget
        form = PartieUpdateForm(request.POST, instance=partie)

        if form.is_valid() and form.has_changed():
            titre = request.POST.get('titre')
            new_budget = form.cleaned_data.get('budget')
            dif = new_budget - old_budget

            if dif > budget_object.solde:
                data["depassement"] = True
                data["message"] = "Le nouveau budget dépasse le budget restant"

            else:
                data["depassement"] = False
                partie.titre = titre
                partie.budget = partie.budget_restant = new_budget
                partie.save()
                message = "Partie modifiée"
                data["message"] = message
                data["valid_form"] = True
                data["partie"] = model_to_dict(partie)

                if dif != 0 and partie.chapitre_set.count() != 0:
                    for chapitre in partie.chapitre_set.all():
                        chapitre.budget = chapitre.budget_restant = chapitre.solde = 0
                        chapitre.save()

                        for article in chapitre.article_set.all():
                            article.budget = 0
                            article.montant = 0
                            article.save()
                    message += ", les budgets des chapitres et articles de cette partie ont été mis à 0. Veuillez les " \
                               "modifier "
                    data["message"] = message
                budget_object.solde -= dif
                budget_object.debit += dif
                budget_object.save()
                data["solde"] = budget_object.solde
        else:
            data["valid_form"] = False
            if not form.has_changed():
                data["changed"] = False
                message = "Aucun changement effectué"
                data["message"] = message
            else:
                data["changed"] = True
                field_errors = {}
                valid_fields = []
                for field in form:
                    if field.errors:
                        field_errors[field.name] = field.errors
                        continue
                    else:
                        valid_fields.append(field.name)

                data["form_valid"] = False
                data["field_errors"] = field_errors
                data["valid_fields"] = valid_fields
        return JsonResponse(data)

    context = {'parties': parties, "solde": budget_object.solde}
    return render(request, "app1/ModifierPartie.html", context)


# SUPPRESSION D'UNE PARTIE #####################################################
def partie_delete_view(request):
    parties = Partie.objects.all()
    year = datetime.datetime.now().year
    budget_object = Budget.objects.filter(date__year=year).first()
    data = dict()

    if request.method == 'POST':
        partie = Partie.objects.get(pk=request.POST.get('partie'))
        data["partie"] = model_to_dict(partie)
        budget_object.solde += partie.budget
        budget_object.save()
        partie.delete()
        data["message"] = "Partie supprimée"
        data["solde"] = budget_object.solde
        data["count"] = Partie.objects.all().count()
        return JsonResponse(data)

    context = {"parties": parties, "solde": budget_object.solde}

    return render(request, 'app1/SupprimerPartie.html', context)

def create_chapitre_ajax(request):
    partie = Partie.objects.get(pk=request.GET.get('partie'))
    data = {"solde": partie.budget_restant}

    return JsonResponse(data)

# AJOUT D'UN CHAPITRE #########################################################
def chapitre_create_view(request):
    parties = Partie.objects.all()
    data = dict()

    if request.method == 'POST':
        form = ChapitreModelForm(request.POST)
        if form.is_valid():
            data["form_valid"] = True
            partie = Partie.objects.get(pk=request.POST.get('partie'))
            titre = form.cleaned_data.get('titre')
            budget = form.cleaned_data.get('budget')
            chapitre = Chapitre.objects.create(titre=titre, partie=partie, budget=budget, solde=budget, budget_restant=budget)
            partie.nbchap += 1
            partie.budget_restant -= budget
            data["solde"] = partie.budget_restant
            partie.save()
            message = 'Chapitre ajouté'
            data["message"] = message
            chapitre.save()
            data["chapitre"] = model_to_dict(chapitre)
        else:
            data["form_valid"] = False
            field_errors = {}
            valid_fields = []

            for field in form:
                if field.errors:
                    field_errors[field.name] = field.errors
                    continue
                else:
                    valid_fields.append(field.name)

            data["field_errors"] = field_errors
            data["valid_fields"] = valid_fields

        return JsonResponse(data)

    context = {"parties": parties}
    return render(request, "app1/AjouterChapitre.html", context)

def chapitre_update_ajax_view(request):
    data = dict()
    if request.GET.get("chapitre") is not None:
        chapitre = Chapitre.objects.get(pk=request.GET.get('chapitre'))
        for partie in Partie.objects.all():
            if chapitre in partie.chapitre_set.all():
                data["partie"] = model_to_dict(partie)
                break
        data["chapitre"] = model_to_dict(chapitre)

    parties = []
    for partie in Partie.objects.all():
        parties.append(model_to_dict(partie))
    data['parties'] = parties
    return JsonResponse(data)

# MISE A JOUR D'UN CHAPITRE #######################################################
def chapitre_update_view(request):
    chapitres = Chapitre.objects.all()
    parties = Partie.objects.all()
    data = dict()

    if request.method == 'POST':
        chapitre = get_object_or_404(Chapitre, pk=request.POST.get("chapitre"))
        form = ChapitreUpdateForm(request.POST, instance=chapitre)
        old_budget = chapitre.budget
        old_partie = chapitre.partie
        new_budget = int(request.POST.get("budget"))

        if form.is_valid() and form.has_changed():
            new_partie = Partie.objects.get(pk=request.POST.get("partie"))

            dif = new_budget - old_budget

            if old_partie.id != new_partie.id:
                if new_budget > new_partie.budget_restant:
                    depassement = True
                    data["depassement"] = True
                    message = "Operation a échoué car le budget de ce chapitre dépasse le budget restant de la partie"

                else:
                    depassement = False
                    new_partie.budget_restant -= new_budget
                    data["solde"] = new_partie.budget_restant
                    new_partie.save()
                    old_partie.budget_restant += old_budget
                    old_partie.save()
                    message = "Chapitre modifiée"
                    if dif != 0:
                        for article in chapitre.article_set.all():
                            article.budget = article.montant = article.solde = 0
                            article.save()
                        message += ", les budgets des articles ont été mis à 0. Veuillez les modifier"
            else:
                if dif > old_partie.budget_restant:
                    depassement = True
                    data["depassement"] = True
                    message = "Operation a échoué car le budget de ce chapitre dépasse le budget restant de la partie"
                else:
                    depassement = False
                    old_partie.budget_restant -= dif
                    old_partie.save()
                    message = "Chapitre modifiée"
                    if dif != 0:
                        for article in chapitre.article_set.all():
                            article.budget = article.montant = article.solde = 0
                            article.save()
                        message += ", les budgets des articles ont été mis à 0. Veuillez les modifier"
                data["solde"] = old_partie.budget_restant

            if not depassement:
                chapitre.titre = request.POST.get('titre')
                chapitre.budget = chapitre.budget_restant = new_budget
                chapitre.partie = new_partie
                chapitre.save()

            data["valid_form"] = True
            data["message"] = message
        else:
            if not form.has_changed():
                data["changed"] = False
                message = "Aucun changement effectué"
                data["message"] = message
            else:
                data["changed"] = True
                field_errors = {}
                valid_fields = []

                for field in form:
                    if field.errors:
                        field_errors[field.name] = field.errors
                        continue
                    else:
                        valid_fields.append(field.name)

                data["field_errors"] = field_errors
                data["valid_fields"] = valid_fields
            data["valid_form"] = False

        return JsonResponse(data)

    context = {"parties": parties, "chapitres": chapitres}
    return render(request, 'app1/ModifierChapitre.html', context)


# SUPPRESSION D'UN CHAPITRE ####################################################
def chapitre_delete_view(request):
    chapitres = Chapitre.objects.all()
    data = dict()

    if request.method == 'POST':
        chapitre = Chapitre.objects.get(pk=request.POST.get('chapitre'))
        data["chapitre"] = model_to_dict(chapitre)
        chapitre.partie.budget_restant += chapitre.budget
        chapitre.partie.nbchap -= 1
        chapitre.partie.save()
        chapitre.delete()

        data["message"] = "Chapitre supprimé"
        data["count"] = Chapitre.objects.all().count()
        return JsonResponse(data)

    context = {"chapitres": chapitres}

    return render(request, 'app1/SupprimerChapitre.html', context)


def article_ajax_create_view(request):
    data = dict()
    chapitre = Chapitre.objects.get(pk=request.GET.get("chapitre"))
    data["solde"] = chapitre.budget_restant

    return JsonResponse(data)

# AJOUT D'UN ARTICLE ###########################################################
def article_create_view(request):
    chapitres = Chapitre.objects.all()
    data = dict()
    count = Article.objects.all().count()

    if request.method == "POST":
        form = ArticleModelForm(request.POST)

        if form.is_valid():
            budget = form.cleaned_data["budget"]
            montant = form.cleaned_data["montant"]
            chapitre = Chapitre.objects.get(pk=request.POST.get("chapitre"))

            if budget < montant:
                data["depassement"] = True

            else:
                article = Article.objects.create(contenu=request.POST.get("contenu"), chapitre=chapitre, montant=montant,
                                       solde=budget, budget=budget, numero_article=count+1)
                chapitre.budget_restant -= budget
                chapitre.nbart += 1
                chapitre.save()
                article.save()
                data["article"] = model_to_dict(article)
            data["solde"] = chapitre.budget_restant
            data["form_valid"] = True
            data["message"] = "Article ajouté"


        else:
            field_errors = {}
            valid_fields = []

            for field in form:
                if field.errors:
                    field_errors[field.name] = field.errors
                    continue
                else:
                    valid_fields.append(field.name)
            data["field_errors"] = field_errors
            data["valid_fields"] = valid_fields
            data["form_valid"] = False

        return JsonResponse(data)

    context = {"chapitres": chapitres}

    return render(request, 'app1/AjouterArticle.html', context)


def ajax_update_article_view(request):
    """"
    data = dict()

    if request.GET.get("article") is not None:
        article = model_to_dict(Article.objects.get(pk=request.GET.get('article')))
        data["article"] = article

    if request.GET.get("chapitre") is not None:
        chapitre = Chapitre.objects.get(pk=request.GET.get('chapitre'))
        data["solde"] = chapitre.budget_restant

    return JsonResponse(data)
    """
    data = dict()
    if request.GET.get("article") is not None:
        article = Article.objects.get(pk=request.GET.get('article'))
        for chapitre in Chapitre.objects.all():
            if article in chapitre.article_set.all():
                data["chapitre"] = model_to_dict(chapitre)
                break
        data["article"] = model_to_dict(article)

    chapitres = []
    for chapitre in Chapitre.objects.all():
        chapitres.append(model_to_dict(chapitre))
    data['chapitres'] = chapitres
    return JsonResponse(data)



# MISE A JOUR D'UN ARTICLE ########################################################
def article_update_view(request):
    articles = Article.objects.all()
    chapitres = Chapitre.objects.all()
    data = dict()

    if request.method == 'POST':
        article = get_object_or_404(Article, pk=request.POST.get("article"))
        form = ArticleModelForm(request.POST, instance=article)
        old_budget = article.budget
        old_chapitre = article.chapitre

        if form.is_valid() and form.has_changed():
            new_chapitre = Chapitre.objects.get(pk=request.POST.get("chapitre"))

            new_budget = int(request.POST.get("budget"))
            dif = new_budget - old_budget

            if old_chapitre.id != new_chapitre.id:
                if new_budget > new_chapitre.budget_restant:
                    data["depassement"] = True
                    message = "Operation a échoué car le budget de cet article dépasse le budget restant de ce chapitre"

                else:
                    new_chapitre.budget_restant -= new_budget
                    data["solde"] = new_chapitre.budget_restant
                    new_chapitre.save()
                    old_chapitre.budget_restant += old_budget
                    old_chapitre.save()
                    message = "Article modifiée"
            else:
                if dif > old_chapitre.budget_restant:
                    data["depassement"] = True
                    message = "Operation a échoué car le budget de cet article dépasse le budget restant de ce chapitre"
                else:
                    old_chapitre.budget_restant -= dif
                    old_chapitre.save()

                    message = "Article modifiée"
                data["solde"] = old_chapitre.budget_restant

            article.titre = request.POST.get('titre')
            article.budget = new_budget
            article.chapitre = new_chapitre
            article.save()
            data["valid_form"] = True
            data["message"] = message
        else:
            if not form.has_changed():
                data["changed"] = False
                message = "Aucun changement effectué"
                data["message"] = message
            else:
                data["changed"] = True
                field_errors = {}
                valid_fields = []

                for field in form:
                    if field.errors:
                        field_errors[field.name] = field.errors
                        continue
                    else:
                        valid_fields.append(field.name)

                data["field_errors"] = field_errors
                data["valid_fields"] = valid_fields
            data["valid_form"] = False

        return JsonResponse(data)


    context = {'articles': articles, "chapitres": chapitres}
    return render(request, 'app1/ModifierArticle.html', context)


# SUPPRESSION D'UN ARTICLE ########################################################
def article_delete_view(request):
    articles = Article.objects.all()
    articles_ = []

    for article in articles:
        demandes = article.demande_set.all()
        if demandes is None or (demande.etat != "non traitée" for demande in demandes):
            articles_.append(article)

    data = dict()

    if request.method == 'POST':
        article = Article.objects.get(pk=request.POST.get('article'))
        data['article'] = model_to_dict(article)

        article.chapitre.budget_restant += article.budget
        article.chapitre.nbart -= 1
        article.chapitre.save()
        article.delete()

        data["message"] = "article supprimé"
        data["count"] = Article.objects.all().count()
        return JsonResponse(data)

    context = {"articles": articles}

    return render(request, 'app1/SupprimerArticle.html', context)


# RECHERCHE DU PROGRAMME ##########################################################
def search_view(request):
    result = []

    #autocomplete
    if "term" in request.GET and request.GET["term"] is not None:
        term = str(request.GET["term"]).strip()
        if term != "":
            articles = Article.objects.filter(contenu__icontains=term).distinct()

            for article in articles:
                result.append(str(article.contenu)[0:100])
            """
            for chapitre in chapitres:
                for article in chapitre.article_set.all():
                    result.append(article.contenu)
            """

        return JsonResponse(result, safe=False)


    else:
        result = []
        data = dict()

        #ajax request
        if request.GET.get("searched") is not None:
            articles = Article.objects.filter(contenu__icontains=request.GET.get('searched'))


            for article in articles:
                result.append(model_to_dict(article))

            data["result"] = result

        else:
            article = Article.objects.filter(contenu__icontains=request.GET.get('value')).first()
            data["article"] = model_to_dict(article)
        return JsonResponse(data)



# PAGE D'ACCEUIL ###########################################################
@login_required(login_url='login')
def employe_view(request):
    parties = Partie.objects.all()
    articles = Article.objects.all()
    now = datetime.datetime.now()
    annonces = Annonce.objects.filter(Q(date_debut__lte=now), Q(date_fin__gte=now))
    tirages = Annonce.objects.filter(date_fin=now)
    non_read = Message.objects.filter(is_read=False).count()
    messages = Message.objects.filter(receiver=request.user)
    context = {"parties": parties, 'articles': articles, "annonces": annonces, "non_read": non_read,
               "messages": messages, "tirages": tirages}

    return render(request, 'app1/Employe.html', context)


# GENERATION DU NOM D'UTILISATEUR ########################################################
def generate_username(email):
    index = str(email).index('@')
    username = str(email)[0: index]
    return username


# GENERATION DU MOT DE PASSE #############################################################
def generate_password():
    letters = string.ascii_letters
    digits = string.digits
    symbols = string.punctuation

    password = ''.join(random.choice(letters + digits + symbols) for _ in range(8))

    return password



# IMPORTATION DU FICHIER EXCEL ##############################################################
def upload_file_view(request):
    file = request.FILES["file"]
    index = str(file).index(".")
    data = dict()

    if str(file)[index + 1:] != "xlsx":
        return HttpResponseBadRequest("error in file format")

    workbook = load_workbook(filename=file)
    sheet = workbook.active
    employees = []

    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue


        employee = Employee.objects.create(nom=row[0], prenom=row[1], date_naissance=row[2], adresse=row[3],
                                           poste=row[4], ccp=row[5], nss=row[6], email=row[7], num_tel=row[8],sex=row[11])

        employee.save()
        username = generate_username(row[7])
        password = generate_password()

        user = User.objects.create_user(username=username, email=row[7], password=password)
        user.save()

        profile = UserProfile.objects.create(user=user, employe=employee,
                                             profile_picture="images/default_image.jpg")
        profile.save()

        employee = model_to_dict(employee)
        employee["photo"] = json.dumps(employee, cls=ExtendedEncoder)
        employees.append(employee)
        data["employees"] = employees

        to = [user.email]

        message = "Hi {}, your password is {}, please click here {} to change your password".format(username, password, "http://127.0.0.1:8000/login")
        email = EmailMessage("Password change", message, "zakimouzaoui123@gmail.com", to)
        EmailThread(email).start()

    return JsonResponse(data)


def admin_crud_view(request):
    fields = []
    employees = Employee.objects.all()

    for field in Employee._meta.fields:
        fields.append(field)

    context = {"fields": fields, 'employees': employees}

    return render(request, 'app1/admin_crud.html', context)


def add_employe_view(request):
    data = dict()

    if request.method == "POST":
        form = EmployeeForm(request.POST)


        if form.is_valid():
            employee = form.save()
            employees = Employee.objects.all()

            email = request.POST.get('email')
            username = generate_username(email)
            password = generate_password()
            user = User.objects.create_user(username=username, email=email, password=password)
            user.save()

            profile = UserProfile.objects.create(user=user, employe=employee,
                                                     profile_picture="images/default_image.jpg")
            profile.save()

            to = [email]

            email = EmailMessage("Password change", "Hi {}, your password is {}, please click here {} to change your password"
                          .format(username, password, "http://127.0.0.1:8000/change_password/{}".format(user.id)),
                          "zakimouzaoui123@gmail.com",
                          to)

            EmailThread(email).start()

            data['form_valid'] = True
            data['html_employee_list'] = render_to_string('app1/employee_list.html', {
                    'employees': employees
            })
            employee = model_to_dict(employee)
            employee["photo"] = json.dumps(employee, cls=ExtendedEncoder)
            data["employee"] = employee
        else:
            data["form_valid"] = False
            field_errors = {}
            field_valid = []
            for field in form:
                if field.errors:
                    for error in field.errors:
                        field_errors[field.name] = error
                    continue
                else:
                    field_valid.append(field.name)

            data["field_errors"] = field_errors
            data["field_valid"] = field_valid

    return JsonResponse(data)


def edit_employee_view(request, employee_id):
    data = dict()
    employee = Employee.objects.get(pk=employee_id)

    if request.method == 'POST':
        form = EmployeeEditForm(request.POST, instance=employee)
        if form.is_valid() and form.has_changed():
            form.save()
            employees = Employee.objects.all()
            html_employee_list = render_to_string("app1/employee_list.html", {"employees": employees}, request=request)
            data["html_employee_list"] = html_employee_list
            data["form_valid"] = True
        else:
            if not form.is_valid():
                data["form_valid"] = False
                field_errors = {}
                field_valid = []
                for field in form:
                    if field.errors:
                        for error in field.errors:
                            field_errors[field.name] = error
                        continue
                    else:
                        field_valid.append(field.name)

                data["field_errors"] = field_errors
                data["field_valid"] = field_valid
            else:
                print("form didn't change")

    employee = model_to_dict(employee)
    employee["photo"] = json.dumps(employee, cls=ExtendedEncoder)
    data["employee"] = employee

    return JsonResponse(data)

# SUPPRESSION D'UN UTILISATEUR ####################################################
def delete_employe_view(request, employee_id):
    data = dict()
    employee = Employee.objects.get(pk=employee_id)

    if request.method == "POST":
        employee.userprofile.user.delete()
        employee.delete()
        data["employee_id"] = employee_id

    else:
        employee = model_to_dict(employee)
        employee["photo"] = json.dumps(employee, cls=ExtendedEncoder)
        data["employee"] = employee

    return JsonResponse(data)

def delete_multiple_employee_view(request):
    data = dict()
    employees = []

    ids = str(request.GET.get('id')).split(',')

    for employee_id in ids:
        employee = Employee.objects.get(id=employee_id)
        employees.append(employee.id)
        employee.userprofile.user.delete()
        employee.delete()

    data["employees"] = employees
    return JsonResponse(data)


def page_not_found_view(request):
    return render(request, 'app1/404.html')


def inscription_view(request):
    if request.method == 'POST':
        return HttpResponse("inscription avec success")
    return render(request, "app1/inscription.html")


def demande_view(request, article_id):
    data = dict()

    if request.method == 'POST':
        form = DemandeForm(request.POST)
        if form.is_valid():
            data["form_valid"] = True

            article = Article.objects.get(id=article_id)
            user = request.user
            if Demande.objects.filter(article=article, user=user).exists():
                data["exists"] = True
            else:
                data["exists"] = False
                demande = form.save(commit=False)
                demande.article = article
                demande.user = user
                comment = request.POST.get("comment")
                paiment = request.POST.get("mode_paiment")
                demande.mode_paiement = paiment

                if comment !="":
                    demande.commentaire = comment
                    for user_ in User.objects.filter(email="zakimouzaoui123@gmail.com"):
                        Message.objects.create(sender=user, receiver=user_, message=comment)
                else:
                    demande.commentaire = "Aucun commentaire"
                demande.save()

                for file in request.FILES.getlist("doc"):
                    f = Fichier.objects.create(document=file, demande=demande)
                    f.save()

        else:
            data["form_valid"] = False
    return JsonResponse(data)


def demande_list_view(request):
    demandes = Demande.objects.all()
    return render(request, "app1/demandes.html", {"demandes": demandes})

def ajax_demande_view(request):
    data = dict()

    demande_id = request.GET.get("demande_id")
    demande = Demande.objects.get(id=demande_id)

    fichiers = []
    for fichier in demande.fichier_set.all():
        fichiers.append(fichier.document.url)

    data["article"] = model_to_dict(demande.article)
    employee = model_to_dict(demande.user.userprofile.employe)
    employee["photo"] = json.dumps(employee, cls=ExtendedEncoder)
    data["employee"] = employee
    data["demande"] = model_to_dict(demande)
    data["fichiers"] = fichiers
    data["mode_paiement"] = demande.mode_paiement

    return JsonResponse(data)

def traiter_demande_view(request):
    data = dict()
    justification = request.POST.get("justificatif")
    demande = Demande.objects.get(id=request.POST.get("demande_id"))

    # cas d'un refus de la demande
    if justification is not None:
        email = EmailMessage("Réponse de la demande", "Bonjour {}, votre demande concernant le chapitre '{}' a été refusée, justification {}"
                  .format(demande.user.username, demande.article.chapitre, justification),
                  "zakimouzaoui123@gmail.com", [demande.user.email])
        EmailThread(email).start()
        user = demande.user
        article = demande.article
        Demande.objects.filter(user=user, article=article).delete()

    # cas d'une acceptation d'une demande
    else:
        form = AccepterDemandeForm(request.POST, instance=demande)
        if form.is_valid() and form.has_changed():
            data["valid"] = True

            if demande.etat != "acceptée":
                montant = demande.article.montant
                changed = False

                if montant <= demande.article.solde:
                    data["suffisant"] = True
                    demande.article.solde -=  montant
                    demande.article.debit +=  montant
                    demande.article.save()
                    demande.etat = "acceptée"
                else:
                    data["suffisant"] = False
            else:
                data["suffisant"] = True
                changed = True

            mode_paiement = request.POST.get("mode_paiement")
            date = request.POST.get("date_paiement")
            demande.mode_paiement = mode_paiement
            demande.date_paiement = date
            demande.save()

            if changed:
                objet = "Demande modifiée"
                message = "Votre demande a été modifiée"
            else:
                objet = "Réponse de la demande"
                message = "Votre demande a été acceptée"

            message = "Bonjour {},".format(demande.user.username) + message + ", mode de paiment:{}, date de paiement:{}".format(
                mode_paiement, date)

            Message.objects.create(sender=request.user, receiver=demande.user, message=message)
            email = EmailMessage(objet, message, "zakimouzaoui123@gmail.com", [demande.user.email])
            EmailThread(email).start()

        else:
            if not form.has_changed():
                data["changed"] = False
            else:
                data["changed"] = True
            data["valid"] = False

    return JsonResponse(data)


def create_pret_view(request):
    data = dict()

    if request.method == 'POST':
        form = PretForm(request.POST)
        if form.is_valid():
            data["form_valid"] = True
            user = request.user
            pret = form.save(commit=False)
            pret.user = user
            mode_paiement = request.POST.get("mode_paiement")
            pret.mode_paiement = mode_paiement
            commentaire = request.POST.get('commentaire')
            if commentaire != "":
                for user_ in User.objects.filter(email="zakimouzaoui123@gmail.com"):
                    Message.objects.create(sender=user, receiver=user_, message=commentaire)
                pret.commentaire = commentaire
            else:
                pret.commentaire = "Aucun commentaire"
            pret.save()

            for file in request.FILES.getlist("doc"):
                f = Fichier_Pret.objects.create(document=file, pret=pret)
                f.save()

        else:
            data["form_valid"] = False
    return JsonResponse(data)


def ajax_pret_view(request):
    data = dict()

    pret_id = request.GET.get("pret_id")
    pret = Pret.objects.get(id=pret_id)

    fichiers = []
    for fichier in pret.fichier_pret_set.all():
        fichiers.append(fichier.document.url)

    employee = model_to_dict(pret.user.userprofile.employe)
    employee["photo"] = json.dumps(employee, cls=ExtendedEncoder)
    data["employee"] = employee
    data["pret"] = model_to_dict(pret)
    data["fichiers"] = fichiers
    data["mode_paiement"] = pret.mode_paiement
    data["commentaire"] = pret.commentaire

    return JsonResponse(data)


def traiter_pret_view(request):
    data = dict()
    justification = request.POST.get("justificatif")
    pret = Pret.objects.get(id=request.POST.get("pret_id"))

    # cas d'un refus de la demande du pret
    if justification is not None:
        message = "Bonjour {}, votre demande du pret avec le montant {} da a été refusée. Justification: {}"\
            .format(pret.user.username, pret.montant, justification)
        Message.objects.create(sender=request.user, receiver=pret.user, message=message)
        email = EmailMessage("Réponse de la demande", message, "zakimouzaoui123@gmail.com", [pret.user.email])
        EmailThread(email).start()
        user = pret.user
        Pret.objects.filter(id=pret.id, user=user).delete()

    # cas d'une acceptation de la demande du pret
    else:
        budget_object = Budget.objects.filter(date__year=datetime.datetime.now().year).first()
        form = AccepterPretForm(request.POST, instance=pret)
        if form.is_valid() and form.has_changed():
            data["valid"] = True
            if pret.etat != "acceptée":
                montant = pret.montant
                changed = False

                if montant <= budget_object.solde:
                    data["suffisant"] = True
                    budget_object.solde -= montant
                    budget_object.debit += montant
                    budget_object.save()
                    pret.etat = "acceptée"
                else:
                    data["suffisant"] = False
            else:
                data["suffisant"] = True
                changed = True

            mode_paiement = request.POST.get("mode_paiement")
            date = request.POST.get("date_paiement")
            pret.mode_paiement = mode_paiement
            pret.date_paiement = date
            pret.save()


            # Partie remboursement
            nombre_mois = int(request.POST.get("nombre_mois"))
            if nombre_mois == 0:
                montant_remobourser = pret.montant
            else:
                montant_remobourser = pret.montant // nombre_mois
                if pret.montant % nombre_mois != 0:
                    nombre_mois += 1

            remboursement = Remboursement.objects.create(pret=pret, nombre_mois=nombre_mois,
                                                         date_debut=date, montant=montant_remobourser)
            remboursement.save()

            if changed:
                objet = "Demande modifiée"
                message_ = "Votre demande a été modifiée "
            else:
                objet = "Réponse de la demande du pret"
                message_ = "Votre demande a été acceptée "

            message = "Bonjour {},".format(pret.user.username) + message_ + "la date de debut de remboursement : {}, le montant de remboursement : {}, " \
                                                                           "le nombre de mois de remboursement : {}, le mode de paiement {}".\
                format(date, remboursement.montant, remboursement.nombre_mois, mode_paiement)
            Message.objects.create(sender=request.user, receiver=pret.user, message=message)
            email = EmailMessage(objet, message, "zakimouzaoui123@gmail.com", [pret.user.email])
            EmailThread(email).start()

        else:
            if not form.has_changed():
                data["changed"] = False
            else:
                data["changed"] = True
            data["valid"] = False

    return JsonResponse(data)

def pret_list_view(request):
    prets = Pret.objects.all()
    return render(request, "app1/Prets.html", {"prets": prets})


def annonce_create_view(request):
    data = dict()
    form = AnnonceForm(request.POST, request.FILES)

    if form.is_valid():
        titre = request.POST.get('titre')
        detail = request.FILES.get('detail')
        date_debut = request.POST.get('date_debut')
        date_fin = request.POST.get('date_fin')
        photo = request.FILES.get('photo')
        budget = request.POST.get('budget')
        annonce = Annonce.objects.create(titre=titre, detail=detail, photo=photo, date_debut=date_debut, date_fin=date_fin, budget=budget)
        annonce.save()
        data["form_valid"] = True
        annonce = model_to_dict(annonce)
        annonce["photo"] = json.dumps(annonce, cls=ExtendedEncoder)
        annonce["detail"] = json.dumps(annonce, cls=ExtendedEncoder)
        data["annonce"] = annonce

    else:
        data["form_valid"] = False
        field_errors = {}
        field_valid = []
        if form.non_field_errors:
            data["non_fields"] = True
        else:
            data["non_fields"] = False

        for field in form:
            if field.errors:
                data["non_fields"] = False
                for error in field.errors:
                    field_errors[field.name] = error
                continue
            else:
                field_valid.append(field.name)

        data["field_errors"] = field_errors
        data["field_valid"] = field_valid

    return JsonResponse(data)


def ajax_update_annonce_view(request):
    data = dict()
    annonce = Annonce.objects.get(id=request.GET.get('annonce_id'))
    data["titre"] = annonce.titre
    data["date_debut"] = annonce.date_debut.date()
    data["date_fin"] = annonce.date_fin.date()
    data["photo"] = annonce.photo.url
    data["detail"] = annonce.detail.url

    return JsonResponse(data)


def annonce_update_view(request):
    data = dict()
    annonce = Annonce.objects.get(id=request.POST.get('annonce'))
    form = UpdateAnnonceForm(request.POST, request.FILES, instance=annonce)
    new_photo = request.FILES.get('photo')
    new_detail = request.FILES.get('detail')

    if not form.is_valid():
        data["valid"] = False
        data["has_changed"] = True
        field_errors = {}
        field_valid = []
        if form.non_field_errors:
            data["non_fields"] = True
        else:
            data["non_fields"] = False

        for field in form:
            if field.errors:
                data["non_fields"] = False
                for error in field.errors:
                    field_errors[field.name] = error
                continue
            else:
                field_valid.append(field.name)

        data["field_errors"] = field_errors
        data["field_valid"] = field_valid

    else:
        data["valid"] = True
        annonce = form.save(commit=False)
        annonce.photo = new_photo
        annonce.detail = new_detail
        annonce.save()
        data["photo"] = annonce.photo.url
        data["detail"] = annonce.detail.url
        annonce = model_to_dict(annonce)
        annonce["photo"] = json.dumps(annonce, cls=ExtendedEncoder)
        annonce["detail"] = json.dumps(annonce, cls=ExtendedEncoder)
        annonce["date_fin"] = annonce["date_fin"].date()
        data["annonce"] = annonce

    return JsonResponse(data)



def annonce_delete_view(request):
    data = dict()

    if request.POST:
        annonce = Annonce.objects.get(id=request.POST.get("annonce"))
        data["id"] = annonce.id
        annonce.delete()

    return JsonResponse(data)


def annonce_inscription_view(request):
    data = dict()
    user = request.user
    annonce = Annonce.objects.get(id=request.GET.get("annonce_id"))

    if request.user in annonce.users.all():
        data["exists"] = True
    else:
        annonce.users.add(user)
        annonce.save()
        Inscription.objects.create(user=user, annonce=annonce)
        data["exists"] = False


    """
    if Annonce.objects.filter(users=user.id, id=req).exists():
        data["exists"] = True
        send_mail("Inscription d'annonce", "Bonjour {}, vous etes inscris à l'annonce {}"
                  .format(user.username, annonce.objet), "zakimouzaoui123@gmail.com", [user.email], fail_silently=False)
                  """

    return JsonResponse(data)


def inscriptions_list_view(request):
    inscriptions = Inscription.objects.all()
    context = {"inscriptions": inscriptions}
    return render(request, "app1/inscriptions.html", context)


def accepter_inscription_view(request):
    data = dict()
    inscription = Inscription.objects.get(id=request.POST.get('inscription_id'))
    annonce = inscription.annonce
    user = inscription.user
    year = datetime.datetime.now().year
    budget_object = Budget.objects.get(date__year=year)
    solde = budget_object.solde
    pourcentage = int(request.POST.get("pourcentage"))
    montant = (100 - pourcentage) * annonce.budget / 100

    if user not in annonce.admis.all():
        data["admis"] = False
        if montant > solde:
            data["suffisant"] = False
        else:
            data["suffisant"] = True
            budget_object.solde -= montant
            budget_object.save()
            inscription.montant = montant
            inscription.save()
            data["solde"] = budget_object.solde
            annonce.admis.add(user)
            objet = "Tirage au sort de l'annonce {}".format(annonce.titre)
            message = "Félicitations! vous etes parmis les gagnants de tirage au sort de l'annonce {}".format(annonce.titre)
            email = request.user.email
            to = [user.email]

            email = EmailMessage(objet, message, email, to)
            EmailThread(email).start()
            m = Message.objects.create(sender=request.user, receiver=user, message=message)
            m.save()
            annonce.save()
    else:
        data["admis"] = True

    return JsonResponse(data)


def ajax_budget_view(request):
    year = datetime.datetime.now().year
    data = dict()
    data["solde"] = Budget.objects.get(date__year=year).solde

    return JsonResponse(data)


def refuser_inscription_view(request):
    data = dict()
    inscription = Inscription.objects.get(id=request.GET.get('inscription_id'))
    annonce = inscription.annonce
    user = inscription.user

    if user in annonce.admis.all():
        year = datetime.datetime.now().year
        budget_object = Budget.objects.get(date__year=year)
        annonce.admis.remove(user)
        montant = inscription.montant
        budget_object.solde += montant
        budget_object.save()
        objet = "Tirage au sort de l'annonce '{}'".format(annonce.titre)
        message = "Malheureusement votre nom n'a pas été tiré au sort au tirage de l'annonce '{}'! Nous vous souhaitons " \
                "beaucoup plus de chance pour le prochain tirage inchallah!".format(annonce.titre)
        email = request.user.email
        to = [user.email]

        email = EmailMessage(objet, message, email, to)
        EmailThread(email).start()
        m = Message.objects.create(sender=request.user, receiver=user, message=message)
        m.save()
        annonce.save()

    return JsonResponse(data)


def pv_create_view(request):
    form = PvModelForm(request.POST, request.FILES)
    data = dict()

    if form.is_valid():
        pv = form.save()
        data["date"] = pv.date
        data["pv"] = pv.pv.url
        data["titre"] = pv.titre
        data["valid"] = True
        data["id"] = pv.id
    else:
        data["valid"] = False

    return JsonResponse(data)


def pv_ajax_view(request):
    pv = Pv.objects.get(id=request.GET.get("pv_id"))
    return JsonResponse({"titre": pv.titre})


def modifier_pv_view(request):
    data = dict()
    pv = Pv.objects.get(id=request.POST.get("pv_id"))

    form = PvModelForm(request.POST, instance=pv)

    if form.is_valid():
        pv = form.save(commit=False)
        pv.pv = request.FILES.get("pv")
        pv.save()
        data["valid"] = True
        data["titre"] = pv.titre
        data["date"] = pv.date
        data["pdf"] = pv.pv.url
        data["id"] = pv.id

    else:
        if not form.has_changed():
            data["changed"] = False

        else:
            data["changed"] = True
        data["valid"] = False
    return JsonResponse(data)


def supprimer_pv_view(request):
    data = dict()

    pv = Pv.objects.get(id=request.POST.get('pv'))
    data["id"] = pv.id
    pv.delete()

    return JsonResponse(data)


def notification_view(request):
    data = dict()
    message = Message.objects.get(id=request.GET.get('message_id'))
    if message.is_read:
        data["read"] = True
    else:
        message.is_read = True
        data["read"] = False
    message.save()
    data["message"] = model_to_dict(message)

    return JsonResponse(data)

def reset_password_view(request):
    if request.method == "POST":
        form = PasswordResetForm(request.POST)

        if form.is_valid():
            email = form.cleaned_data.get('email')
            for user in User.objects.filter(email=email):
                subject = "Password Reset"
                parameters = {
                    "email": email,
                    "domain": "http://127.0.0.1:8000",
                    "site_name": "Horizon",
                    "uid": urlsafe_base64_encode(force_bytes(user.id)),
                    "token": default_token_generator.make_token(user),
                    "protocol": "http"
                }
                html_message = render_to_string("app1/password_reset_message.html", context=parameters)
                email = EmailMessage(subject, html_message, '', [user.email])
                EmailThread(email).start()
            return redirect("password_reset_done")

    return render(request, "app1/reset_password.html")


def password_reset_done_view(request):
    return render(request, "app1/message.html")


class ResetPasswordConfirmView(PasswordResetConfirmView):
    template_name = "app1/reset_password_confirm.html"


def change_password_view(request):
    if request.user.is_authenticated:
        if request.method == "POST":
            data = dict()
            form = PasswordChangeForm(user=request.user, data=request.POST)

            if form.is_valid():
                data["valid"] = True
                new_password = form.cleaned_data["new_password1"]
                request.user.set_password(new_password)
                request.user.save()
            else:
                data["valid"] = False
                if form["old_password"].errors:
                    data["old_password_error"] = True
                if form["new_password2"].errors:
                    data["new_password2_error"] = True
            return JsonResponse(data)
    else:
        return redirect("login")

    return render(request, "app1/change_password.html")


def change_image_view(request):
    data = dict()
    image_file = request.FILES.get("image_file")
    request.user.userprofile.profile_picture = image_file
    request.user.userprofile.save()
    data["profile_picture"] = request.user.userprofile.profile_picture.url
    return JsonResponse(data)


def user_profile_view(request):
    return render(request, "app1/page profile.html")


def ajouter_budget_view(request):
    date = request.POST.get('date')
    year = int(date.split('-')[0])
    data = dict()

    if Budget.objects.filter(date__year=year).exists():
        data["exists"] = True
    else:
        data["exists"] = False
        montant = request.POST.get('montant')
        Budget.objects.create(date=date, montant=montant, solde=montant, debit=0, credit=0)

    return JsonResponse(data)


def ajouter_recette_view(request):
    data = dict()
    montant = int(request.POST.get('montant'))
    date = request.POST.get('date')
    subvention = request.POST.get('subvention')
    year = int(date.split('-')[0])

    if not Budget.objects.filter(date__year=year).exists():
        data["exists"] = False
    else:
        data["exists"] = True
        budget = Budget.objects.get(date__year=year)
        budget.solde += montant
        budget.credit += montant
        budget.save()
        Recette.objects.create(date=date, subvention=subvention, montant=montant)

    return JsonResponse({})


# serializer les images et les fichiers
class ExtendedEncoder(DjangoJSONEncoder):
    def default(self, o):
        if isinstance(o, ImageFieldFile):
            return str(o)
        elif isinstance(o, InMemoryUploadedFile):
            return str(o)
        else:
            return super().default(o)

# My love *-*
class EmailThread(threading.Thread):
    def __init__(self, email):
        self.email = email
        threading.Thread.__init__(self)

    def run(self):
        self.email.send(fail_silently=False)
